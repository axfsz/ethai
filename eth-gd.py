import requests
import pandas as pd
import time
from datetime import datetime
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# 获取币安 ETHUSDT 最新价格和多时间框架K线数据
def get_binance_data():
    try:
        # 获取最新价格
        ticker_url = "https://api.binance.com/api/v3/ticker/price?symbol=ETHUSDT"
        price_resp = requests.get(ticker_url).json()
        current_price = float(price_resp['price'])

        # 根据新要求，获取4h, 1h, 15m三个时间框架的K线数据
        # 4小时数据，用于判断主趋势
        kline_url_4h = "https://api.binance.com/api/v3/klines?symbol=ETHUSDT&interval=4h&limit=50"
        klines_4h = requests.get(kline_url_4h).json()

        # 1小时数据，用于识别波段推进结构
        kline_url_1h = "https://api.binance.com/api/v3/klines?symbol=ETHUSDT&interval=1h&limit=100"
        klines_1h = requests.get(kline_url_1h).json()

        # 15分钟数据，用于精确入场
        kline_url_15m = "https://api.binance.com/api/v3/klines?symbol=ETHUSDT&interval=15m&limit=50"
        klines_15m = requests.get(kline_url_15m).json()

        return {
            "current_price": current_price,
            "klines_4h": klines_4h,
            "klines_1h": klines_1h,
            "klines_15m": klines_15m
        }

    except Exception as e:
        logging.error(f"获取多时间框架数据失败: {e}")
        return None

# --- 辅助计算函数 ---
def calc_ema(data, period):
    if len(data) < period:
        return np.array([np.mean(data)])
    ema = [sum(data[:period]) / period]
    alpha = 2 / (period + 1)
    for price in data[period:]:
        ema.append(alpha * price + (1 - alpha) * ema[-1])
    return np.array(ema)

# --- 缠论核心分析函数 ---

def find_segments(strokes):
    """从笔构建线段（简化版）"""
    if not strokes or len(strokes) < 3:
        return []
    segments = []
    i = 0
    while i < len(strokes) - 2:
        # 三笔同向，合并为一段
        if strokes[i]['type'] == strokes[i+1]['type'] == strokes[i+2]['type']:
            segment_type = strokes[i]['type']
            segment = {
                'start_price': strokes[i]['start_price'],
                'end_price': strokes[i+2]['end_price'],
                'start_time': strokes[i]['start_time'],
                'end_time': strokes[i+2]['end_time'],
                'type': segment_type
            }
            segments.append(segment)
            i += 2
        else:
            i += 1
    return segments

def find_fractals(klines):
    """识别分型"""
    fractals = []
    # 忽略最新K线，因为它尚未走完
    for i in range(2, len(klines) - 2):
        highs = [float(k[2]) for k in klines[i-2:i+3]]
        lows = [float(k[3]) for k in klines[i-2:i+3]]

        # 顶分型: 中间K线的最高点是5根中最高的
        if highs[2] == max(highs):
            fractals.append({'index': i, 'type': 'top', 'price': highs[2], 'time': klines[i][0]})
        
        # 底分型: 中间K线的最低点是5根中最低的
        if lows[2] == min(lows):
            fractals.append({'index': i, 'type': 'bottom', 'price': lows[2], 'time': klines[i][0]})
            
    return fractals

def find_strokes(fractals):
    """从分型构建笔"""
    if not fractals:
        return []

    strokes = []
    last_fractal = fractals[0]

    for i in range(1, len(fractals)):
        current_fractal = fractals[i]
        
        # 确保顶底分型交替出现
        if current_fractal['type'] == last_fractal['type']:
            # 如果连续出现同类型分型，保留价格更高(顶)或更低(底)的那个
            if current_fractal['type'] == 'top' and current_fractal['price'] > last_fractal['price']:
                last_fractal = current_fractal
            elif current_fractal['type'] == 'bottom' and current_fractal['price'] < last_fractal['price']:
                last_fractal = current_fractal
        else:
            # 出现交替的分型，形成一笔
            # 检查K线数量是否大于等于5根 (缠论要求)
            if abs(current_fractal['index'] - last_fractal['index']) >= 4:
                stroke_type = 'up' if last_fractal['type'] == 'bottom' else 'down'
                strokes.append({
                    'start_price': last_fractal['price'],
                    'end_price': current_fractal['price'],
                    'start_time': last_fractal['time'],
                    'end_time': current_fractal['time'],
                    'type': stroke_type
                })
                last_fractal = current_fractal

    return strokes

def create_order(name, direction, trigger, order_price, sl, tp, investment, leverage, analysis):
    quantity_formula = f'={investment}/{order_price}'
    if direction == "BUY":
        profit_formula = f'=({tp}-{order_price})*G2*I2' # Assuming G2 is quantity, I2 is leverage
        loss_formula = f'=({order_price}-{sl})*G2*I2'
    else:
        profit_formula = f'=({order_price}-{tp})*G3*I3' # Assuming G3 is quantity, I3 is leverage
        loss_formula = f'=({sl}-{order_price})*G3*I3'

    return {
        "策略名称": name,
        "方向": direction,
        "触发信号": trigger,
        "挂单价格": order_price,
        "止损价格": sl,
        "止盈价格": tp,
        "数量": quantity_formula,
        "杠杆倍数": leverage,
        "预计盈利": profit_formula,
        "预计亏损": loss_formula,
        "策略分析": analysis
    }

# 动态生成挂单表 (V4 - 基于缠论思想)
def generate_order_table(market_data, investment_amount, leverage):
    current_price = market_data['current_price']
    orders = []
    fractals_15m = find_fractals(market_data['klines_15m'])
    strokes_15m = find_strokes(fractals_15m)
    segments_15m = find_segments(strokes_15m)
    zhongshus_15m = find_zhongshu(segments_15m)
    buy1_15m, sell1_15m = find_first_second_buy_sell(strokes_15m, zhongshus_15m)
    fractals_1h = find_fractals(market_data['klines_1h'])
    strokes_1h = find_strokes(fractals_1h)
    segments_1h = find_segments(strokes_1h)
    zhongshus_1h = find_zhongshu(segments_1h)
    buy1_1h, sell1_1h = find_first_second_buy_sell(strokes_1h, zhongshus_1h)
    # 生成中枢买卖点策略
    if buy1_15m:
        analysis = f"15M中枢{buy1_15m['type']}信号，突破中枢区间[{buy1_15m['zs']['start']:.2f},{buy1_15m['zs']['end']:.2f}]，入场价{buy1_15m['price']:.2f}"
        orders.append(create_order("ETH缠论中枢多单", "BUY", current_price, round(buy1_15m['price'],2), round(buy1_15m['zs']['start'],2), round(buy1_15m['price']+abs(buy1_15m['price']-buy1_15m['zs']['start'])*2,2), investment_amount, leverage, analysis))
    if sell1_15m:
        analysis = f"15M中枢{sell1_15m['type']}信号，跌破中枢区间[{sell1_15m['zs']['start']:.2f},{sell1_15m['zs']['end']:.2f}]，入场价{sell1_15m['price']:.2f}"
        orders.append(create_order("ETH缠论中枢空单", "SELL", current_price, round(sell1_15m['price'],2), round(sell1_15m['zs']['end'],2), round(sell1_15m['price']-abs(sell1_15m['price']-sell1_15m['zs']['end'])*2,2), investment_amount, leverage, analysis))
    # ... existing code ...
def find_zhongshu(segments):
    """识别中枢（简化版，三段重叠区间）"""
    if not segments or len(segments) < 3:
        return []
    zhongshus = []
    for i in range(len(segments)-2):
        s1, s2, s3 = segments[i], segments[i+1], segments[i+2]
        # 取三段的重叠区间
        zs_high = min(s1['end_price'], s2['end_price'], s3['end_price'])
        zs_low = max(s1['start_price'], s2['start_price'], s3['start_price'])
        if zs_high > zs_low:
            zhongshus.append({'start':zs_low, 'end':zs_high, 'index':i+2})
    return zhongshus

def find_first_second_buy_sell(strokes, zhongshus):
    """识别第一、第二类买卖点（简化版）"""
    if not zhongshus:
        return None, None
    zs = zhongshus[-1]
    # 第一类买点：离开中枢后一笔向上突破中枢上沿
    if strokes[-1]['type']=='up' and strokes[-1]['end_price']>zs['end'] and strokes[-2]['end_price']<=zs['end']:
        return {'type':'first_buy','price':strokes[-1]['end_price'],'zs':zs}, None
    # 第一类卖点：离开中枢后一笔向下跌破中枢下沿
    if strokes[-1]['type']=='down' and strokes[-1]['end_price']<zs['start'] and strokes[-2]['end_price']>=zs['start']:
        return None, {'type':'first_sell','price':strokes[-1]['end_price'],'zs':zs}
    # 第二类买点：回踩中枢上沿后再次向上
    if strokes[-2]['type']=='up' and strokes[-1]['type']=='down' and strokes[-1]['end_price']>zs['end'] and strokes[-1]['start_price']>=zs['end']:
        return {'type':'second_buy','price':strokes[-1]['end_price'],'zs':zs}, None
    # 第二类卖点：反抽中枢下沿后再次向下
    if strokes[-2]['type']=='down' and strokes[-1]['type']=='up' and strokes[-1]['end_price']<zs['start'] and strokes[-1]['start_price']<=zs['start']:
        return None, {'type':'second_sell','price':strokes[-1]['end_price'],'zs':zs}
    return None, None

    def find_last_buy_signal(strokes):
        if len(strokes) < 4 or not (strokes[-1]['type'] == 'down' and strokes[-3]['type'] == 'down'):
            return None
        
        last_down_stroke = strokes[-1]
        prev_down_stroke = strokes[-3]

        # 简化背驰判断：当前下跌笔的幅度小于前一个下跌笔
        if (last_down_stroke['start_price'] - last_down_stroke['end_price']) < (prev_down_stroke['start_price'] - prev_down_stroke['end_price']):
            # 确保价格没有创新低
            if last_down_stroke['end_price'] > prev_down_stroke['end_price']:
                 return {'price': last_down_stroke['end_price'], 'stop_loss': prev_down_stroke['end_price']}
        return None

    def find_last_sell_signal(strokes):
        if len(strokes) < 4 or not (strokes[-1]['type'] == 'up' and strokes[-3]['type'] == 'up'):
            return None

        last_up_stroke = strokes[-1]
        prev_up_stroke = strokes[-3]

        if (last_up_stroke['end_price'] - last_up_stroke['start_price']) < (prev_up_stroke['end_price'] - prev_up_stroke['start_price']):
            if last_up_stroke['end_price'] < prev_up_stroke['end_price']:
                return {'price': last_up_stroke['end_price'], 'stop_loss': prev_up_stroke['end_price']}
        return None

    # --- 3. 生成策略 --- 
    buy_signal_15m = find_last_buy_signal(strokes_15m)
    sell_signal_15m = find_last_sell_signal(strokes_15m)
    buy_signal_1h = find_last_buy_signal(strokes_1h)
    sell_signal_1h = find_last_sell_signal(strokes_1h)

    # 长期单：1h和15m信号共振
    if buy_signal_1h and buy_signal_15m:
        order_price = round(current_price * 1.001, 2)
        stop_loss = round(buy_signal_1h['stop_loss'] * 0.99, 2)
        risk = order_price - stop_loss
        take_profit = round(order_price + risk * 3, 2)
        analysis = f"长期看多信号(1H与15M共振):\n1H级别出现买入结构，止损参考位: {buy_signal_1h['stop_loss']:.2f}。\n15M级别出现买入结构，入场点参考: {buy_signal_15m['price']:.2f}。\n策略: 等待价格回调至入场点附近买入，严格止损。"
        orders.append(create_order("ETH缠论长线多单", "BUY", current_price, order_price, stop_loss, take_profit, investment_amount, leverage, analysis))

    if sell_signal_1h and sell_signal_15m:
        order_price = round(current_price * 0.999, 2)
        stop_loss = round(sell_signal_1h['stop_loss'] * 1.01, 2)
        risk = stop_loss - order_price
        take_profit = round(order_price - risk * 3, 2)
        analysis = f"长期看空信号(1H与15M共振):\n1H级别出现卖出结构，止损参考位: {sell_signal_1h['stop_loss']:.2f}。\n15M级别出现卖出结构，入场点参考: {sell_signal_15m['price']:.2f}。\n策略: 等待价格反弹至入场点附近卖出，严格止损。"
        orders.append(create_order("ETH缠论长线空单", "SELL", current_price, order_price, stop_loss, take_profit, investment_amount, leverage, analysis))

    # 短期单：仅15m信号
    elif buy_signal_15m:
        order_price = round(current_price * 1.001, 2)
        stop_loss = round(buy_signal_15m['stop_loss'] * 0.995, 2)
        risk = order_price - stop_loss
        take_profit = round(order_price + risk * 1.5, 2) # 短期目标放低
        analysis = f"短期看多信号(15M):\n15M级别出现潜在底背驰买入结构。\n入场点参考: {buy_signal_15m['price']:.2f}，止损参考: {buy_signal_15m['stop_loss']:.2f}。\n策略: 短线操作，快进快出，盈亏比目标1.5:1。"
        orders.append(create_order("ETH缠论短线多单", "BUY", current_price, order_price, stop_loss, take_profit, investment_amount, leverage, analysis))

    elif sell_signal_15m:
        order_price = round(current_price * 0.999, 2)
        stop_loss = round(sell_signal_15m['stop_loss'] * 1.005, 2)
        risk = stop_loss - order_price
        take_profit = round(order_price - risk * 1.5, 2)
        analysis = f"短期看空信号(15M):\n15M级别出现潜在顶背驰卖出结构。\n入场点参考: {sell_signal_15m['price']:.2f}，止损参考: {sell_signal_15m['stop_loss']:.2f}。\n策略: 短线操作，快进快出，盈亏比目标1.5:1。"
        orders.append(create_order("ETH缠论短线空单", "SELL", current_price, order_price, stop_loss, take_profit, investment_amount, leverage, analysis))

    return pd.DataFrame(orders)



last_status_notify_time = 0

# 主任务：获取数据、生成订单并写入Excel
def run_main_task():
    global last_status_notify_time
    output_file = "ETH_动态挂单表.xlsx"
    investment_amount = 5000  # 投资资金
    leverage = 10  # 杠杆倍数

    logging.info("正在获取实时市场数据...")
    data = get_binance_data()
    if data:
        logging.info(f"当前价格: {data['current_price']}")

        df = generate_order_table(data, investment_amount, leverage)

        if df.empty:
            logging.info("当前无满足条件的交易信号，不生成新表。")
            # 每小时通知一次无信号状态
            current_time = time.time()
            if current_time - last_status_notify_time > 3600:
                try:
                    requests.post("http://localhost:5001/notify_status")
                    logging.info("已发送无交易机会的状态通知。")
                    last_status_notify_time = current_time
                except Exception as e:
                    logging.error(f"发送状态通知失败: {e}")
            return

        try:
            book = openpyxl.load_workbook(output_file)
            while len(book.sheetnames) >= 2:
                oldest_sheet_name = book.sheetnames[0]
                if oldest_sheet_name != 'Init':
                    book.remove(book[oldest_sheet_name])
                    logging.info(f"🗑️ 已删除最旧的工作表: {oldest_sheet_name}")
                else: # 如果最旧的是Init，则删除下一个
                    if len(book.sheetnames) > 1:
                        book.remove(book[book.sheetnames[1]])
                        logging.info(f"🗑️ 已删除最旧的工作表: {book.sheetnames[1]}")
            book.save(output_file)
        except FileNotFoundError:
            pass

        sheet_name = datetime.now().strftime('%Y-%m-%d %H_%M')
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width
        logging.info(f"📈 新的挂单表已生成: {sheet_name}")
        # 触发策略通知
        try:
            requests.post("http://localhost:5001/notify_order_strategy", params={"file": output_file})
            logging.info("已触发策略通知")
        except Exception as e:
            logging.error(f"触发策略通知失败: {e}")

# 主程序入口
if __name__ == "__main__":
    output_file = "ETH_动态挂单表.xlsx"
    try:
        book = openpyxl.load_workbook(output_file)
        if not book.sheetnames or 'Init' not in book.sheetnames:
            book.create_sheet("Init", 0)
            book.save(output_file)
    except FileNotFoundError:
        book = openpyxl.Workbook()
        book.create_sheet("Init", 0)
        book.save(output_file)
    
    logging.info("✅ 实时监控已启动，每15分钟执行一次策略分析。")

    while True:
        try:
            run_main_task()
            logging.info("本次任务执行完毕，等待15分钟后再次运行...")
            time.sleep(900)  # 等待15分钟
        except KeyboardInterrupt:
            logging.info("程序被用户中断。")
            break
        except Exception as e:
            logging.error(f"程序发生未知错误: {e}")
            logging.info("发生错误，等待5分钟后重试...")
            time.sleep(300) # 发生错误时，等待5分钟再重试
